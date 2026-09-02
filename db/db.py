from sqlalchemy import create_engine, Integer, Text
from sqlalchemy.orm import DeclarativeBase, Mapped, mapped_column, Session
from openpyxl import load_workbook

# ---------------------- 数据库连接 ----------------------
engine = create_engine("sqlite:///game.db", echo=False)


# ---------------------- 实体基类 ----------------------
class Base(DeclarativeBase):
    pass


# ---------------------- 实体类定义 ----------------------
class Item(Base):
    __tablename__ = "t_item"

    id: Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
    item_id: Mapped[str] = mapped_column(Text, nullable=False)
    item_name: Mapped[str | None] = mapped_column(Text, nullable=True)
    item_desc: Mapped[str | None] = mapped_column(Text, nullable=True)
    item_type: Mapped[int | None] = mapped_column(Integer, nullable=True)
    need_level: Mapped[int | None] = mapped_column(Integer, nullable=True)
    quality: Mapped[int | None] = mapped_column(Integer, nullable=True)

    def __repr__(self):
        return f"<Item(id={self.id},item_id={self.item_id},item_name={self.item_name})>"


class Task(Base):
    __tablename__ = "t_task"

    id: Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
    task_id: Mapped[str] = mapped_column(Text, nullable=False)
    task_name: Mapped[str | None] = mapped_column(Text, nullable=True)
    task_desc: Mapped[str | None] = mapped_column(Text, nullable=True)
    task_type: Mapped[int | None] = mapped_column(Integer, nullable=True)
    need_item_ids: Mapped[str | None] = mapped_column(Text, nullable=True)
    need_item_count: Mapped[str | None] = mapped_column(Text, nullable=True)
    need_money: Mapped[int | None] = mapped_column(Integer, nullable=True)
    reward_item_ids: Mapped[str | None] = mapped_column(Text, nullable=True)
    reward_item_counts: Mapped[str | None] = mapped_column(Text, nullable=True)
    reward_wellknow: Mapped[int | None] = mapped_column(Integer, nullable=True)
    reward_exp: Mapped[int | None] = mapped_column(Integer, nullable=True)
    reward_money: Mapped[int | None] = mapped_column(Integer, nullable=True)

    def __repr__(self):
        return f"<Task(id={self.id},task_id={self.task_id},task_name={self.task_name})>"


class Nobility(Base):
    __tablename__ = "t_nobility"

    id: Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
    name: Mapped[str] = mapped_column(Text, nullable=False)
    well_know: Mapped[int | None] = mapped_column(Integer, nullable=True)

    def __repr__(self):
        return f"<Nobility(id={self.id},name={self.name},well_know={self.well_know})>"


class Business(Base):
    __tablename__ = "t_business"

    id: Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
    level: Mapped[int | None] = mapped_column(Integer, nullable=True)
    task_name: Mapped[str | None] = mapped_column(Text, nullable=True)
    item_name: Mapped[str | None] = mapped_column(Text, nullable=True)
    item_count: Mapped[str | None] = mapped_column(Text, nullable=True)

    def __repr__(self):
        return f"<Business(id={self.id},level={self.level})>"


class BusinessLevel(Base):
    __tablename__ = "t_business_level"

    id: Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
    level: Mapped[int | None] = mapped_column(Integer, nullable=True)
    exp: Mapped[int | None] = mapped_column(Integer, nullable=True)
    money: Mapped[int | None] = mapped_column(Integer, nullable=True)
    item_ids: Mapped[str | None] = mapped_column(Text, nullable=True)
    item_counts: Mapped[str | None] = mapped_column(Text, nullable=True)
    need_order_num: Mapped[int | None] = mapped_column(Integer, nullable=True)

    def __repr__(self):
        return f"<BusinessLevel(id={self.id},level={self.level})>"


# 创建全部数据表（不存在才创建）
Base.metadata.create_all(engine)


# ---------------------- 通用CRUD工具函数 ----------------------
# 新增
def add_one(model_obj):
    with Session(engine) as session:
        session.add(model_obj)
        session.commit()
        session.refresh(model_obj)
        print(f"新增成功，主键id={model_obj.id}")
        return model_obj


# 根据id查询单条
def get_by_id(model_cls, uid: int):
    with Session(engine) as session:
        res = session.get(model_cls, uid)
        if not res:
            print(f"{model_cls.__tablename__} 未找到id={uid}")
        return res


# 查询全部
def get_all(model_cls):
    with Session(engine) as session:
        data_list = session.query(model_cls).all()
        for item in data_list:
            print(item)
        return data_list

def load_item():
    try:
        with open("../res/itemInfo.txt", "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                row = line.split("ˇ")
                item_type = int(row[7])
                quality = 0
                if item_type == 12 or item_type == 13 or item_type == 14:
                    quality = 1
                item = Item(item_id=row[3], item_name=row[4], item_desc=row[2], item_type=item_type, need_level=int(row[8]), quality=quality)
                add_one(item)
    except FileNotFoundError:
        print(f"文件不存")
    except Exception as e:
        print(f"解析失败：{e}")

def load_task():
    try:
        with open("../res/taskInfo.txt", "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                row = line.split("ˇ")
                task = Task(task_id=row[0], task_name=row[10], task_desc=row[1], task_type=int(row[8]), need_item_ids=row[5],
                    need_item_count=row[6], need_money=int(row[15]), reward_item_ids=row[3], reward_item_counts=row[4],
                    reward_wellknow=int(row[7]), reward_exp=int(row[11]), reward_money=int(row[2]))
                add_one(task)
    except FileNotFoundError:
        print(f"文件不存")
    except Exception as e:
        print(f"解析失败：{e}")

def load_nobility():
    datas = [
        ("乡绅",200),
        ("绅士",500),
        ("骑士",800),
        ("二等男爵",1200),
        ("男爵",1500),
        ("二等子爵",2100),
        ("子爵",5300),
        ("二等伯爵",8500),
        ("伯爵",13600),
        ("二等侯爵",21800),
        ("侯爵",35000),
        ("二等公爵",56200),
        ("公爵",90200),
        ("终身公爵",144800),
        ("世袭公爵",232600),
        ("大公爵",367800),
        ("边境亲王",574000),
        ("宫廷亲王",879200),
        ("辅国亲王",1320500),
        ("大亲王",1944600)
    ]
    for data in datas:
        nobility = Nobility(name=data[0], well_know=data[1])
        add_one(nobility)

def replace_segment(str):
    return str.replace("，", ",").replace("；", ",").replace("、", ",")

# update t_business set item_name='杰诺瓦士蛋糕' where item_name='杰诺瓦斯蛋糕';
# update t_business set item_name='杰诺瓦士蛋糕' where item_name='杰若瓦士蛋糕';
# update t_business set item_name='火鸡三明治' where item_name='鸡腿三明治';
# update t_business set item_name='贵族塞米隆' where item_name='贵族塞米龙';
# update t_business set item_name='西蒙塔尔牛肉' where item_name='西门塔尔牛肉';
# update t_business set item_name='生菜色拉' where item_name='生菜沙拉';
# update t_business set item_name='皮埃蒙特牛肉' where item_name='皮尔蒙特牛肉';
# update t_business set item_name='伊比利亚猪肉' where item_name='伊比利黑猪';
# update t_business set item_name='火鸡肉' where item_name='黑火鸡肉';
# update t_business set item_name='水果色拉' where item_name='水果沙拉';
# update t_business set item_name='弗郎萨克玫瑰红葡萄酒' where item_name='弗郎萨克玫瑰红葡萄酒';
# update t_business set item_name='葡汁烤羊排' where item_name='葡汗烤羊排';
# update t_business set item_name='意式蔬菜色拉' where item_name='意式蔬菜沙拉';
# update t_business set item_name='西红柿酱' where item_name='番茄酱';
# update t_business set item_name='石料' where item_name='石板';
# update t_business set item_name='无花果慕斯蛋糕' where item_name='无花果幕斯蛋糕';
# update t_business set item_name='单峰驼' where item_name='单峰骆驼';
# update t_business set item_name='佛罗伦萨烧兔肉' where item_name='佛罗伦萨烤兔肉';
# update t_business set item_name='黑加仑子酒' where item_name='黑加仑酒';
# update t_business set item_name='佛罗伦萨烧兔肉' where item_name='弗洛伦撒烧兔肉';
# update t_business set item_name='奶油野菇汁兔肉卷' where item_name='奶油草菇汁兔肉卷';
# update t_business set item_name='鸡肉芝士脆卷' where item_name='鸡肉芝士脆卷 ';
# update t_business set item_name='无花果慕斯蛋糕' where item_name='无花果幕斯';
# update t_business set item_name='烤乳鸽填栗子馅' where item_name='烤乳鸽填栗子陷';
# update t_business set item_name='博斯沃思奶酪' where item_name='博斯沃斯奶酪';
def load_business():
    wb = load_workbook("../res/商会数据表.xlsx")  # 加载整个工作簿
    ws = wb["成长任务"]  # 指定工作表
    # 遍历所有行
    for row in ws.iter_rows(min_row=2, max_row=194, values_only=True):
        level = int(row[0])
        task_name = row[1]
        item_names = replace_segment(str(row[3])).split(",")
        item_counts = replace_segment(str(row[4])).split(",")
        if len(item_names) == len(item_counts):
            for index, value in enumerate(item_names):
                business = Business(level=level, task_name=task_name, item_name=value,item_count=item_counts[index])
                add_one(business)
        else:
            print(task_name + ":物品数量有误")

def load_business_level():
    wb = load_workbook("../res/商会数据表.xlsx")  # 加载整个工作簿
    ws = wb["升级任务"]  # 指定工作表
    # 遍历所有行
    for row in ws.iter_rows(min_row=3, max_row=17, values_only=True):
        level = int(row[0])
        exp = int(row[1])
        money = int(row[2])
        item_ids = replace_segment(str(row[3]))
        item_counts = replace_segment(str(row[4]))
        need_order_num = int(row[5])
        business_level = BusinessLevel(level=level, exp=exp, money=money, item_ids=item_ids, item_counts=item_counts, need_order_num=need_order_num)
        add_one(business_level)
        add_business_level_item(item_ids, item_counts, level)

def add_business_level_item(item_ids, item_counts, level):
    items = item_ids.split(",")
    counts = item_counts.split(",")
    if len(items) == len(counts):
        for index, value in enumerate(items):
            business = Business(level=level, task_name="升级任务", item_name=value, item_count=counts[index])
            add_one(business)
    else:
        print(str(level) + "级任务物品数量有误")

# ---------------------- 测试示例 ----------------------
if __name__ == "__main__":
    # load_item()
    # load_task()
    # load_nobility()
    # load_business()
    # load_business_level()
    print(111)
